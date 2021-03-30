#!/usr/local/bin/python
# -*- coding: utf-8 -*-
import requests 
import json 
import io 
import os 
import time 
import datetime 
import pyodbc 
import concurrent.futures 
import openpyxl 
import pandas
from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select


options = Options()
chrome_path = "C:\\PYTHON\\driver\\chromedriver" 
tmap = int(time.time())
localtime = time.strftime("%m%d",time.localtime())
#print(localtime)
#driver = webdriver.Chrome(chrome_path,options=options) 

#id = 1 第一個品牌
#http://www.hantao888.com/brand.php?id=1
#driver.get("http://www.hantao888.com/brand.php?id=1")
#driver.find_element_by_css_selector('.head_nav_list > li:nth-child(2)').click()
#driver.find_element_by_css_selector('.filter_brand > a:nth-child(1)').click()
##a = driver.find_element_by_css_selector('div.good_box:nth-child(1) > p:nth-child(2)').get_attribute('value')
#b = driver.find_element_by_css_selector('div.good_box:nth-child(1) > p:nth-child(3) > font:nth-child(1)').get_attribute('text')
#print(a)
#print(b)
def filterbrand(filter_brand):
    if filter_brand[0] == '=' :
        #print (filter_brand[0])
        return str(filter_brand[1])
    else:
        #print (filter_brand)
        return str(filter_brand)



def GetApi():
    u = "http://www.hantao888.com/brand.php"
    resp = requests.get(u)
    resp.encoding = ("utf-8")
    html = resp.text
    #print(html)
    AAA = BeautifulSoup(html)
    filter_brand = AAA.select("div.filter_brand a")
    NO = len(filter_brand)
    print(NO)
    try :
        for q in range(0,NO):
            filter_brand = str(AAA.select("div.filter_brand a")[q]["href"][-2:])
            filter_brandTEST = AAA.select("div.filter_brand a")[q]["title"]
            print(filterbrand(filter_brand)) 
            print(filter_brandTEST)
            
        
            header = {
                "Referer":	"http://www.hantao888.com/mobile/brand.php?id="+str(q),
                'Cache-control': 'private',
                'Connection': 'keep-alive',
                'Content-Encoding': 'gzip',
                'Content-Type': 'text/html; charset=utf-8',
            }           
            data = {
                'ast':'0',
                'amount':'10',
                
            }
            qq = str(q)
            #print(qq)``
            url = 'http://www.hantao888.com/mobile/brand.php?act=asynclist&category=0&brand='+filterbrand(filter_brand)+'&price_min=&price_max=&filter_attr=&page=1&sort=last_update&order=DESC'
            urlS = 'http://www.hantao888.com/mobile/brand.php?id='+filterbrand(filter_brand)
            
            print(urlS)
            resp2 = requests.post(url,headers=header,data=data).text
            respS = requests.post(urlS)
            
        #print(respS)
            #s = json.loads(respS)
            a = json.loads(resp2) 
            P = len(a)
            print(P)
            
            for i in range (0,P) :

                b = a[i]
                c = b['pro-inner']

                #Ssoup = BeautifulSoup(respS.text)
                #status = Ssoup.select("a.btn2")
                #print(len(status))
                #print(Ssoup)
                #print(status)
        
                #抓取產品名稱
                soup = BeautifulSoup(c)
                proTitle = soup.select("div.proTitle a")[0]
                
                #print(title.string)
                #print(PID)
                #抓取產品金額
                proPrice = soup.select("div.proPrice span")[0]
                #抓取照片URL
                imgA = str(soup.select("img")[0]['src'])
                imgurl = 'http://www.hantao888.com' + imgA
                #print(imgurl)
                #img = requests.get(imgurl)
                #print(Getimg)

                
                
                
                #html = requests.get(item.get('src'))   # get函式獲取圖片連結地址，requests傳送訪問請求
                #img_name = folder_path + str(index + 1) +'.png'
                #with open(img_name, 'wb') as file:  # 以byte形式將圖片資料寫入
                #    file.write(html.content)
                #    file.flush()
                #    file.close()  # 關閉檔案
                PorID = str(soup.select("div.proTitle a")[0])
                PID = str(PorID[22:26])
                header= {
                    "Accept":"*/*",
                    "Accept-Encoding":	"gzip, deflate",
                    "Accept-Language":	"zh-TW,zh;q=0.8,en-US;q=0.5,en;q=0.3",
                    "Accept-Language":"zh-TW,zh;q=0.8,en-US;q=0.5,en;q=0.3",
                    "Connection":	"keep-alive",
                    "Cookie":"ECS[display]=grid; zh_choose=t; ECS[history]=1030%2C3194%2C2107%2C636; ECS[visit_times]=9; zh_choose=t; real_ipd=1.200.75.180; ECS_ID=9461df93025d671d8bc7538cb61ea5a466153dac",
                    "Referer":	"http://www.hantao888.com/goods.php?id="+PID,
                    "User-Agent":"Mozilla/5.0 (iPhone; CPU iPhone OS 12_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/12.0 Mobile/15A372 Safari/604.1",
                    }
                #print(header['Referer'])
                if PID[3] == '"' : 
                    #print(PID[0:3])
                    uu = "http://www.hantao888.com/mobile/goods.php?id="+PID
                    PURL = "http://www.hantao888.com/goods.php?act=price&id="+PID[0:3]+"&attr=&number=1&1616771455463463="
                    #print(PURL)
                    prodata = requests.get(uu,headers=header).text 
                    #Prodata = json.loads(prodata)
                    Prodatasoup = BeautifulSoup(prodata)
                    prodata = Prodatasoup.select("a.btn-popupSKU-addcart")[0].text
                    protext = Prodatasoup.findAll("p").text
                    print(protext)
                    getdata = requests.get(PURL,headers=header).text 
                    Getdata = json.loads(getdata)
                    
                    #print(filter_brand,proTitle.string,Getdata["result"][6:-7],Getdata["result1"][6:-7])
                    #print(filter_brand)
                    #print(proTitle.string) 
                    G = str(Getdata["result"][6:-7])
                    GG = str(Getdata["result1"][6:-7])
                    
                    
                    Edata = [PID[0:3],proTitle.string,imgurl,G,GG,prodata]
                
                    #print(Edata)
                    
                else:
                    #print(PID)
                    uu = "http://www.hantao888.com/mobile/goods.php?id="+PID
                    PURL = "http://www.hantao888.com/goods.php?act=price&id="+PID+"&attr=&number=1&1616771455463463="
                    
                    prodata = requests.get(uu,headers=header).text 
                    #Prodata = json.loads(prodata)
                    Prodatasoup = BeautifulSoup(prodata)
                    prodata = Prodatasoup.select("a.btn-popupSKU-addcart")[0].text
                    #protext = Prodatasoup.select("div.desc p")
                    #print(prodata)
                    
                    getdata = requests.get(PURL,headers=header).text   
                    Getdata = json.loads(getdata)   
                    #print(Getdata)       
                    #print(filter_brand,proTitle.string,Getdata["result"][6:-7],Getdata["result1"][6:-7])
                    #print(filter_brand)
                    #print(proTitle.string)
                    G = str(Getdata["result"][6:-7])
                    GG = str(Getdata["result1"][6:-7])
                    Edata = [PID,proTitle.string,imgurl,G,GG,prodata]
                    #print(Edata)
                    #print(Edata)
                #wb = openpyxl.load_workbook('EcTestCase.xlsx')
                #sheet = wb["P"]
                #titles = ("品牌名稱","商品名稱","商品照片URL","價格","價格2")
                #sheet.append(titles)
                #sheet.append(Edata)
                #wb.save("EcTestCase.xlsx")
    except:
        pass        
                
                
                
                
                #with open("C:\\GitHub_projects\\GetMakeupData\\images" + input_img + "str(proTitle.string)" + ".jpg", "wb") as file:  # 開啟資料夾及命名圖片檔
                    #file.write(img.content)  # 寫入圖片的二進位碼
                #print(img)
                #print(imgurl)
                #EXdata = [proTitle.string,imgurl,proPrice.string]
                


        
    #print(b.find_all("div","class=proTitle"))



def MakeExcel(EXdata):
    #print(EXdata)
    Day = str(localtime)
    #print(AddExcelData.EXdata)
    try :
        
        wb = openpyxl.load_workbook('EcTestCase.xlsx')
        sheet = wb[Day]
        sheet.append(EXdata)
    except:
        
        #wb = openpyxl.Workbook()
        wb = openpyxl.load_workbook('EcTestCase.xlsx')
        print(EXdata)
        sheet = wb.create_sheet(Day , 0)
        titles = ("品牌名稱&商品名稱","商品照片URL","價格")
        sheet.append(titles)
        sheet.append(EXdata)
    wb.save("EcTestCase.xlsx")
    
GetApi()
        
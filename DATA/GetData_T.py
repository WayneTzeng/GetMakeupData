#!/usr/local/bin/python
# -*- coding: utf-8 -*-
import requests 
import json 
import io 
import os 
import time 
import datetime 
import pyodbc 
import concurrent.futures 
import openpyxl 
import pandas
from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select


options = Options()
chrome_path = "C:\\PYTHON\\driver\\chromedriver" 
tmap = int(time.time())
localtime = time.strftime("%m%d",time.localtime())
print(localtime)
driver = webdriver.Chrome(chrome_path,options=options) 

#id = 1 第一個品牌
#http://www.hantao888.com/brand.php?id=1
driver.get("http://taohan.kr/?language=FT&p=all_fenlei&list_id=all_goods&brand_id=82")
a = driver.page_source
print(a)
driver.find_element_by_css_selector('.head_nav_list > li:nth-child(2)').click()
driver.find_element_by_css_selector('.filter_brand > a:nth-child(1)').click()
a = driver.find_element_by_css_selector('div.good_box:nth-child(1) > p:nth-child(2)').get_attribute('value')
b = driver.find_element_by_css_selector('div.good_box:nth-child(1) > p:nth-child(3) > font:nth-child(1)').get_attribute('text')
print(a)
print(b)






def GetApi():

    #titles = ("品牌名稱ID","商品名稱","LOGO","商品照片URL","價格","重量","參考介紹")
    #sheet.append(titles)
    INGH = "http://taohan.kr"
    header = {
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "Cookie": "PHPSESSID=rp8mkb2clio2ds2f7lub8cf882; isfirstvisited=false",
        } 
    url = 'http://taohan.kr/?language=FT&p=all_fenlei&list_id=all_goods'
    #print(url)
    resp = requests.get(url)
    resp.encoding = ("utf-8")
    html = resp.text
    #print(html)
    
    soup = BeautifulSoup(html)
    filter_brand = len(soup.select("div.filter_brand a"))
    #print(soup)
    #print(filter_brand)
    for i in range(0,filter_brand):
        brand_id = soup.select("div.filter_brand a")[i]["href"][-2:]
        title = soup.select("div.filter_brand a")[i]["title"]
        src = soup.select("div.filter_brand img")[i]["src"]
        #print(brand_id,title,src)
        Purl = "http://taohan.kr/?language=FT&p=all_fenlei&list_id=all_goods&brand_id="+ brand_id
        #print(Purl)
        resp2 = requests.get(Purl)
        resp2.encoding = ("utf-8")
        html2 = resp2.text
        Psuop = BeautifulSoup(html2)
        #print(Psuop)
        for ii in range(0,len(Psuop.select("div.good_box"))):
            #print(len(Psuop.select("div.good_box")))
            #print(ii)
            #product_ID = Psuop.select("div.good_box a")[ii]["href"][-4:]
            product_name = Psuop.select("p.shop_list_name")[ii].string
            product_src = Psuop.select("div.good_box img")[ii]["src"]
            product_price = Psuop.select("p.shop_list_price")[ii].string[2:-4]
            product_weight = Psuop.select("p.shop_list_weight")[ii].string[5:-4]
            
            #print(product_ID)
            print(product_name)
            #print(product_src)
            #print(product_price)
            #print(product_weight)
            Pdata = [title,product_name,INGH+src,INGH+product_src,product_price,product_weight]
            #Pur3 = "http://taohan.kr/?language=&p=good&id="+ product_ID
            #print(Purl)
            #resp3 = requests.get(Pur3)
            #resp3.encoding = ("utf-8")
            #html3 = resp3.text
            #PsuopInfo = BeautifulSoup(html3)
            #print(PsuopInfo)
            #textdata = []
            
            #for o in range(0,len(PsuopInfo.select("p"))):
            #    good_text = PsuopInfo.select("p")[o].string
            #    if good_text != None :
            #        
            #         textdata += [good_text]
            #     else:
            #         pass
            #print(textdata)
            
            wb = openpyxl.load_workbook('EcTestCase.xlsx')
            sheet = wb["P"]
            sheet.append(Pdata)
            
            #MakeExcel(EXdata)
            wb.save("EcTestCase.xlsx")


#print(b.find_all("div","class=proTitle"))



def MakeExcel(EXdata):
    #print(EXdata)
    Day = str(localtime)
    #print(AddExcelData.EXdata)
    try :
        
        wb = openpyxl.load_workbook('EcTestCase.xlsx')
        sheet = wb[Day]
        sheet.append(EXdata)
    except:
        
        #wb = openpyxl.Workbook()
        wb = openpyxl.load_workbook('EcTestCase.xlsx')
        print(EXdata)
        sheet = wb.create_sheet(Day , 0)
        titles = ("品牌名稱&商品名稱","商品照片URL","價格")
        sheet.append(titles)
        sheet.append(EXdata)
    wb.save("EcTestCase.xlsx")
    
GetApi()
        
#!/usr/local/bin/python
# -*- coding: utf-8 -*-
import requests 
import json 
import io 
import os 
import time 
import datetime 
import pyodbc 
import concurrent.futures 
import openpyxl 
import pandas
from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select


options = Options()
chrome_path = "C:\\PYTHON\\driver\\chromedriver" 
tmap = int(time.time())
localtime = time.strftime("%m%d",time.localtime())
#print(localtime)
#driver = webdriver.Chrome(chrome_path,options=options) 

#id = 1 第一個品牌
#http://www.hantao888.com/brand.php?id=1
#driver.get("http://www.hantao888.com/brand.php?id=1")
#driver.find_element_by_css_selector('.head_nav_list > li:nth-child(2)').click()
#driver.find_element_by_css_selector('.filter_brand > a:nth-child(1)').click()
##a = driver.find_element_by_css_selector('div.good_box:nth-child(1) > p:nth-child(2)').get_attribute('value')
#b = driver.find_element_by_css_selector('div.good_box:nth-child(1) > p:nth-child(3) > font:nth-child(1)').get_attribute('text')
#print(a)
#print(b)






def GetApi():
    for q in range(1,62):
        try :
            header = {
                "Referer":	"http://www.hantao888.com/mobile/brand.php?id="+str(q),
                'Cache-control': 'private',
                'Connection': 'keep-alive',
                'Content-Encoding': 'gzip',
                'Content-Type': 'text/html; charset=utf-8',
            }           
            data = {
                'ast':'0',
                'amount':'10',
                
                
            }
            qq = str(q)
            url = 'http://www.hantao888.com/mobile/brand.php?act=asynclist&category=0&brand='+qq+'&price_min=&price_max=&filter_attr=&page=1&sort=last_update&order=DESC'
            #print(url)
            resp2 = requests.post(url,headers=header,data=data).text
            #print(resp2)
            a = json.loads(resp2)
            P = len(a)
            #print(P)
            for i in range (0,P) :
                b = a[i]
                c = b['pro-inner']
                soup = BeautifulSoup(c)
                #抓取產品名稱
                proTitle = soup.select("div.proTitle a")[0]
                #抓取產品金額
                proPrice = soup.select("div.proPrice span")[0]
                #抓取照片URL
                imgA = str(soup.select("img")[0]['src'])
                imgurl = 'http://www.hantao888.com' + imgA
                print(imgurl)
                img = requests.get(imgurl)
                print(Getimg)
                
                html = requests.get(item.get('src'))   # get函式獲取圖片連結地址，requests傳送訪問請求
                img_name = folder_path + str(index + 1) +'.png'
                with open(img_name, 'wb') as file:  # 以byte形式將圖片資料寫入
                    file.write(html.content)
                    file.flush()
                    file.close()  # 關閉檔案
    
                
                
                
                
                with open("C:\\GitHub_projects\\GetMakeupData\\images" + input_img + "str(proTitle.string)" + ".jpg", "wb") as file:  # 開啟資料夾及命名圖片檔
                    file.write(img.content)  # 寫入圖片的二進位碼
                #print(img)
                #print(imgurl)
                #EXdata = [proTitle.string,imgurl,proPrice.string]
                #print(EXdata)
                #wb = openpyxl.load_workbook('EcTestCase.xlsx')
                #sheet = wb.create_sheet(q)
                #sheet = wb["P"]
                #titles = ("品牌名稱&商品名稱","商品照片URL","價格")
                #sheet.append(titles)
                #sheet.append(EXdata)
                
                #MakeExcel(EXdata)
                #wb.save("EcTestCase.xlsx")
        except:
            pass
        
    #print(b.find_all("div","class=proTitle"))



def MakeExcel(EXdata):
    #print(EXdata)
    Day = str(localtime)
    #print(AddExcelData.EXdata)
    try :
        
        wb = openpyxl.load_workbook('EcTestCase.xlsx')
        sheet = wb[Day]
        sheet.append(EXdata)
    except:
        
        #wb = openpyxl.Workbook()
        wb = openpyxl.load_workbook('EcTestCase.xlsx')
        print(EXdata)
        sheet = wb.create_sheet(Day , 0)
        titles = ("品牌名稱&商品名稱","商品照片URL","價格")
        sheet.append(titles)
        sheet.append(EXdata)
    wb.save("EcTestCase.xlsx")
    
GetApi()
        